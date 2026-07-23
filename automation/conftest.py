import os
import time
import traceback
from datetime import datetime
from pathlib import Path

import pytest
from appium import webdriver
from appium.options.android import UiAutomator2Options
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import sys
ROOT = Path(__file__).resolve().parents[1]
AUTOMATION_ROOT = ROOT / "automation"
if str(AUTOMATION_ROOT) not in sys.path:
    sys.path.insert(0, str(AUTOMATION_ROOT))

REPORT_DIR = AUTOMATION_ROOT / "reports"
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
REPORT_PATH = REPORT_DIR / f"Test_Report_{RUN_TIMESTAMP}.xlsx"
SCREENSHOTS_DIR = REPORT_DIR / "screenshots"
SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)

MODULES = [
    "Login",
    "Registration",
    "Forgot Password",
    "Home",
    "AI Myth Checker",
    "Learning Articles",
    "Learning Videos",
    "Quiz",
    "Brushing Tracker",
    "Notifications",
    "Reports",
    "Profile",
    "Family Members",
    "Settings",
    "Logout",
]


@pytest.fixture(scope="session")
def session_results():
    return []


@pytest.fixture(scope="session")
def appium_driver(request):
    desired_caps = {
        "platformName": "Android",
        "platformVersion": "17",
        "deviceName": "emulator-5554",
        "automationName": "UiAutomator2",
        "app": str(ROOT / "dentzy" / "build" / "app" / "outputs" / "flutter-apk" / "app-debug.apk"),
        "appPackage": "com.example.dentzy",
        "appActivity": "com.example.dentzy.MainActivity",
        "noReset": False,
        "newCommandTimeout": 600,
        "autoGrantPermissions": True,
    }

    try:
        options = UiAutomator2Options().load_capabilities(desired_caps)
        try:
            # Try Appium v2 default URL first
            driver = webdriver.Remote("http://127.0.0.1:4723", options=options)
        except Exception:
            # Fallback to Appium v1 /wd/hub URL
            driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", options=options)
        time.sleep(3)
        return driver
    except Exception as exc:
        print(f"Failed to connect to Appium server: {exc}")
        return None


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    if rep.when == "call":
        item._dentzy_makereport_status = rep.outcome
        if rep.failed:
            item._dentzy_makereport_error = str(rep.longrepr)

@pytest.fixture(autouse=True)
def reset_app_state(appium_driver):
    if appium_driver:
        try:
            appium_driver.terminate_app("com.example.dentzy")
            time.sleep(1)
            appium_driver.activate_app("com.example.dentzy")
            time.sleep(2)
        except:
            pass

@pytest.fixture(autouse=True)
def collect_test_result(request, appium_driver, session_results):
    test_id = request.node.get_closest_marker("test_id")
    if test_id is None:
        test_id = request.node.name
    request.node._dentzy_test_id = test_id.args[0] if hasattr(test_id, "args") and test_id.args else test_id
    request.node._dentzy_module = _get_marker_value(request.node, "module", "General")
    request.node._dentzy_feature = _get_marker_value(request.node, "feature", "General")
    request.node._dentzy_scenario = _get_marker_value(request.node, "scenario", request.node.name)
    request.node._dentzy_test_type = _get_marker_value(request.node, "test_type", "Positive")
    request.node._dentzy_expected = _get_marker_value(request.node, "expected", "Expected behavior")
    request.node._dentzy_result = "Not executed"
    request.node._dentzy_error = ""
    request.node._dentzy_stack = ""
    request.node._dentzy_screenshot = ""
    request.node._dentzy_status = "Skipped"
    request.node._dentzy_execution_time = 0.0
    request.node._dentzy_timestamp = datetime.now().isoformat(timespec="seconds")
    request.node._dentzy_driver = appium_driver

    start = time.perf_counter()
    try:
        if appium_driver is None:
            raise pytest.skip("Appium or emulator unavailable")
        yield
    finally:
        request.node._dentzy_execution_time = round(time.perf_counter() - start, 2)
        
        # Read the outcome from makereport hook
        status = getattr(request.node, "_dentzy_makereport_status", "passed")
        
        if status == "passed":
            request.node._dentzy_status = "Pass"
            request.node._dentzy_result = "Observed expected behavior"
        elif status == "failed":
            request.node._dentzy_status = "Fail"
            request.node._dentzy_result = "Observed unexpected behavior"
            request.node._dentzy_error = getattr(request.node, "_dentzy_makereport_error", "Unknown error")
            request.node._dentzy_screenshot = take_screenshot(request.node)
        else:
            request.node._dentzy_status = "Skipped"
            request.node._dentzy_result = "Test was skipped"
            
        row_data = {
            "Test Case ID": request.node._dentzy_test_id,
            "Module": request.node._dentzy_module,
            "Feature": request.node._dentzy_feature,
            "Test Scenario": request.node._dentzy_scenario,
            "Test Type": request.node._dentzy_test_type,
            "Expected Result": request.node._dentzy_expected,
            "Actual Result": request.node._dentzy_result,
            "Status": request.node._dentzy_status,
            "Execution Time": f"{request.node._dentzy_execution_time}s",
            "Timestamp": request.node._dentzy_timestamp,
            "Screenshot Path": request.node._dentzy_screenshot,
            "Error Message": request.node._dentzy_error,
            "Remarks": "",
        }
        session_results.append(row_data)
        append_to_master_report(row_data)

import os
from openpyxl import load_workbook

MASTER_REPORT = REPORT_DIR / "Final_Test_Report.xlsx"

def append_to_master_report(row_data):
    headers = [
        "Test Case ID", "Module", "Feature", "Test Scenario", "Test Type",
        "Expected Result", "Actual Result", "Status", "Execution Time",
        "Timestamp", "Screenshot Path", "Error Message", "Remarks"
    ]
    try:
        if MASTER_REPORT.exists():
            wb = load_workbook(MASTER_REPORT)
            ws = wb["Test Execution Report"]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Execution Report"
            ws.append(headers)
        
        ws.append([row_data[h] for h in headers])
        wb.save(MASTER_REPORT)
    except Exception as e:
        print(f"Failed to append to master report: {e}")

def _get_marker_value(node, name, default):
    marker = node.get_closest_marker(name)
    if marker is None or not marker.args:
        return default
    return marker.args[0]


def take_screenshot(node):
    driver = getattr(node, "_dentzy_driver", None)
    if driver is None:
        return ""
    try:
        screenshot_path = SCREENSHOTS_DIR / f"{node.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        driver.save_screenshot(str(screenshot_path))
        return str(screenshot_path)
    except Exception:
        return ""


def pytest_configure(config):
    config.addinivalue_line("markers", "test_id: custom test id")
    config.addinivalue_line("markers", "module: module name")
    config.addinivalue_line("markers", "feature: feature name")
    config.addinivalue_line("markers", "scenario: scenario description")
    config.addinivalue_line("markers", "test_type: positive/negative/boundary")
    config.addinivalue_line("markers", "expected: expected result")


def pytest_sessionfinish(session, exitstatus):
    create_report(session)


def create_report(session):
    rows = []
    for item in getattr(session, "items", []):
        if hasattr(item, "_dentzy_test_id"):
            rows.append({
                "Test Case ID": getattr(item, "_dentzy_test_id", item.name),
                "Module": getattr(item, "_dentzy_module", "General"),
                "Feature": getattr(item, "_dentzy_feature", "General"),
                "Test Scenario": getattr(item, "_dentzy_scenario", item.name),
                "Test Type": getattr(item, "_dentzy_test_type", "Positive"),
                "Expected Result": getattr(item, "_dentzy_expected", "Expected behavior"),
                "Actual Result": getattr(item, "_dentzy_result", "Not executed"),
                "Status": getattr(item, "_dentzy_status", "Skipped"),
                "Execution Time": f"{getattr(item, '_dentzy_execution_time', 0)}s",
                "Timestamp": getattr(item, "_dentzy_timestamp", datetime.now().isoformat(timespec="seconds")),
                "Screenshot Path": getattr(item, "_dentzy_screenshot", ""),
                "Error Message": getattr(item, "_dentzy_error", ""),
                "Remarks": "",
            })
    if not rows:
        rows = [{
            "Test Case ID": "N/A",
            "Module": "General",
            "Feature": "General",
            "Test Scenario": "No tests executed",
            "Test Type": "Positive",
            "Expected Result": "No tests executed",
            "Actual Result": "No tests executed",
            "Status": "Skipped",
            "Execution Time": "0s",
            "Timestamp": datetime.now().isoformat(timespec="seconds"),
            "Screenshot Path": "",
            "Error Message": "",
            "Remarks": "",
        }]

    wb = Workbook()
    ws_exec = wb.active
    ws_exec.title = "Test Execution Report"
    headers = [
        "Test Case ID",
        "Module",
        "Feature",
        "Test Scenario",
        "Test Type",
        "Expected Result",
        "Actual Result",
        "Status",
        "Execution Time",
        "Timestamp",
        "Screenshot Path",
        "Error Message",
        "Remarks",
    ]
    ws_exec.append(headers)
    for row in rows:
        ws_exec.append([row[h] for h in headers])

    ws_summary = wb.create_sheet("Summary")
    ws_module = wb.create_sheet("Module Summary")
    ws_failed = wb.create_sheet("Failed Test Cases")

    total = len(rows)
    passed = sum(1 for r in rows if r["Status"] == "Pass")
    failed = sum(1 for r in rows if r["Status"] == "Fail")
    skipped = sum(1 for r in rows if r["Status"] == "Skipped")
    executed = total - skipped
    pass_pct = round((passed / executed * 100), 2) if executed else 0
    fail_pct = round((failed / executed * 100), 2) if executed else 0
    overall_status = "Pass" if failed == 0 and total > 0 and skipped == 0 else "Fail" if failed > 0 else "Skipped"

    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Test Cases", total])
    ws_summary.append(["Executed", executed])
    ws_summary.append(["Passed", passed])
    ws_summary.append(["Failed", failed])
    ws_summary.append(["Skipped", skipped])
    ws_summary.append(["Pass Percentage", f"{pass_pct}%"])
    ws_summary.append(["Fail Percentage", f"{fail_pct}%"])
    ws_summary.append(["Total Execution Time", "N/A"])
    ws_summary.append(["Overall Status", overall_status])
    if failed == 0:
        ws_summary.append(["Note", "All test cases passed successfully."])

    ws_module.append(["Module Name", "Total Test Cases", "Passed", "Failed", "Pass Percentage"])
    seen_modules = set()
    for row in rows:
        seen_modules.add(row["Module"])
    for module in MODULES + sorted([m for m in seen_modules if m not in MODULES and m != "General"]):
        mod_rows = [r for r in rows if r["Module"] == module]
        if not mod_rows:
            continue
        total_mod = len(mod_rows)
        passed_mod = sum(1 for r in mod_rows if r["Status"] == "Pass")
        failed_mod = sum(1 for r in mod_rows if r["Status"] == "Fail")
        pass_pct_mod = round((passed_mod / total_mod * 100), 2) if total_mod else 0
        ws_module.append([module, total_mod, passed_mod, failed_mod, f"{pass_pct_mod}%"])

    ws_failed.append(["Test Case ID", "Module", "Scenario", "Failure Reason", "Screenshot", "Stack Trace"])
    for row in rows:
        if row["Status"] == "Fail":
            ws_failed.append([row["Test Case ID"], row["Module"], row["Test Scenario"], row["Error Message"], row["Screenshot Path"], row["Error Message"]])

    format_workbook(wb)
    wb.save(REPORT_PATH)


def format_workbook(wb):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    pass_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    fail_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    skip_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    header_font = Font(bold=True, color="FFFFFF")

    for sheet_name in ["Test Execution Report", "Summary", "Module Summary", "Failed Test Cases"]:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column in ws.columns:
            max_len = 0
            for cell in column:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)
        ws.auto_filter.ref = ws.dimensions

    ws_exec = wb["Test Execution Report"]
    for row in ws_exec.iter_rows(min_row=2, max_row=ws_exec.max_row):
        status = row[7].value if row[7].value is not None else ""
        fill = pass_fill if status == "Pass" else fail_fill if status == "Fail" else skip_fill if status == "Skipped" else None
        if fill is not None:
            for cell in row:
                cell.fill = fill
